"""Integration tests for the Bulk Import/Export skill (Skill 35).

Tests cover CSV, JSON, and Excel import/export round-trips.
"""

from __future__ import annotations

import csv
import io
import json

from lauren import (
    Bytes,
    LaurenFactory,
    Query,
    QueryField,
    Scope,
    controller,
    get,
    injectable,
    module,
    post,
)
from lauren.exceptions import HTTPError
from lauren.testing import TestClient


class BadRequestError(HTTPError):
    """400 Bad Request — used by bulk import validation."""

    status_code = 400
    code = "bad_request"


# ---------------------------------------------------------------------------
# Domain
# ---------------------------------------------------------------------------


@injectable(scope=Scope.SINGLETON)
class BulkProcessor:
    def import_csv(self, data: bytes, encoding: str = "utf-8") -> list[dict]:
        reader = csv.DictReader(io.StringIO(data.decode(encoding)))
        return list(reader)

    def export_csv(self, records: list[dict]) -> bytes:
        if not records:
            return b""
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(records[0].keys()))
        writer.writeheader()
        writer.writerows(records)
        return buf.getvalue().encode("utf-8")

    def import_json(self, data: bytes) -> list[dict]:
        return json.loads(data)

    def export_json(self, records: list[dict]) -> bytes:
        return json.dumps(records, indent=2).encode("utf-8")

    def import_excel(self, data: bytes) -> list[dict]:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(data))
            ws = wb.active
            rows = list(ws.rows)
            if not rows:
                return []
            headers = [cell.value for cell in rows[0]]
            return [dict(zip(headers, [c.value for c in row])) for row in rows[1:]]
        except ImportError:
            return []

    def export_excel(self, records: list[dict]) -> bytes:
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            if records:
                ws.append(list(records[0].keys()))
                for r in records:
                    ws.append(list(r.values()))
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except ImportError:
            return b""


# ---------------------------------------------------------------------------
# Controller & Module
# ---------------------------------------------------------------------------


@controller("/bulk")
class BulkController:
    def __init__(self, processor: BulkProcessor) -> None:
        self._processor = processor

    @post("/import")
    async def import_data(
        self, body: Bytes, fmt: Query[str] = QueryField(default="csv")
    ) -> dict:
        if fmt == "csv":
            records = self._processor.import_csv(body)
        elif fmt == "json":
            records = self._processor.import_json(body)
        elif fmt == "excel":
            records = self._processor.import_excel(body)
        else:
            raise BadRequestError(f"Unknown format: {fmt}")
        return {"count": len(records), "records": records}

    @get("/export/csv")
    async def export_csv(self) -> dict:
        records = [{"id": "1", "name": "Alice"}, {"id": "2", "name": "Bob"}]
        data = self._processor.export_csv(records)
        return {"csv": data.decode("utf-8"), "count": len(records)}

    @get("/export/json")
    async def export_json(self) -> dict:
        records = [{"id": "1", "name": "Alice"}, {"id": "2", "name": "Bob"}]
        data = self._processor.export_json(records)
        # Return as parsed JSON for easy assertions
        return {"records": json.loads(data), "count": len(records)}

    @get("/export/excel")
    async def export_excel(self) -> dict:
        records = [{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}]
        data = self._processor.export_excel(records)
        return {"size": len(data)}


@module(controllers=[BulkController], providers=[BulkProcessor])
class BulkModule:
    pass


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def build_app() -> TestClient:
    return TestClient(LaurenFactory.create(BulkModule))


def _make_csv(records: list[dict]) -> bytes:
    buf = io.StringIO()
    if not records:
        return b""
    writer = csv.DictWriter(buf, fieldnames=list(records[0].keys()))
    writer.writeheader()
    writer.writerows(records)
    return buf.getvalue().encode("utf-8")


def _make_excel(records: list[dict]) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    if records:
        ws.append(list(records[0].keys()))
        for r in records:
            ws.append(list(r.values()))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestBulkProcessorUnit:
    def test_csv_round_trip(self) -> None:
        processor = BulkProcessor()
        records = [{"name": "Alice", "age": "30"}, {"name": "Bob", "age": "25"}]
        exported = processor.export_csv(records)
        imported = processor.import_csv(exported)
        assert len(imported) == 2
        assert imported[0]["name"] == "Alice"
        assert imported[1]["name"] == "Bob"

    def test_json_round_trip(self) -> None:
        processor = BulkProcessor()
        records = [{"id": 1, "value": "x"}, {"id": 2, "value": "y"}]
        exported = processor.export_json(records)
        imported = processor.import_json(exported)
        assert imported == records

    def test_excel_round_trip(self) -> None:
        processor = BulkProcessor()
        records = [{"name": "Alice", "score": 95}, {"name": "Bob", "score": 88}]
        exported = processor.export_excel(records)
        assert len(exported) > 0
        imported = processor.import_excel(exported)
        assert len(imported) == 2
        assert imported[0]["name"] == "Alice"

    def test_export_csv_empty(self) -> None:
        processor = BulkProcessor()
        assert processor.export_csv([]) == b""

    def test_import_csv_parses_headers(self) -> None:
        processor = BulkProcessor()
        csv_bytes = b"id,name\n1,Alice\n2,Bob\n"
        records = processor.import_csv(csv_bytes)
        assert records[0]["id"] == "1"
        assert records[1]["name"] == "Bob"


class TestBulkController:
    def test_import_csv_via_endpoint(self) -> None:
        client = build_app()
        csv_data = _make_csv(
            [{"name": "Alice", "score": "95"}, {"name": "Bob", "score": "88"}]
        )
        r = client.post("/bulk/import?fmt=csv", content=csv_data)
        assert r.status_code == 200
        body = r.json()
        assert body["count"] == 2
        assert body["records"][0]["name"] == "Alice"

    def test_import_json_via_endpoint(self) -> None:
        client = build_app()
        json_data = json.dumps([{"id": 1, "name": "Alice"}]).encode()
        r = client.post("/bulk/import?fmt=json", content=json_data)
        assert r.status_code == 200
        assert r.json()["count"] == 1

    def test_import_excel_via_endpoint(self) -> None:
        client = build_app()
        excel_data = _make_excel([{"name": "Alice", "score": 95}])
        r = client.post("/bulk/import?fmt=excel", content=excel_data)
        assert r.status_code == 200
        assert r.json()["count"] == 1

    def test_export_csv_endpoint(self) -> None:
        client = build_app()
        r = client.get("/bulk/export/csv")
        assert r.status_code == 200
        body = r.json()
        assert "Alice" in body["csv"]
        assert body["count"] == 2

    def test_export_json_endpoint(self) -> None:
        client = build_app()
        r = client.get("/bulk/export/json")
        assert r.status_code == 200
        body = r.json()
        assert body["count"] == 2
        names = [rec["name"] for rec in body["records"]]
        assert "Alice" in names

    def test_export_excel_endpoint(self) -> None:
        client = build_app()
        r = client.get("/bulk/export/excel")
        assert r.status_code == 200
        assert r.json()["size"] > 0

    def test_import_unknown_format_returns_400(self) -> None:
        client = build_app()
        r = client.post("/bulk/import?fmt=xml", content=b"<root/>")
        assert r.status_code == 400
